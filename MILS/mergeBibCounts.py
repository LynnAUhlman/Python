import os
from openpyxl import Workbook, load_workbook

def merge_workbooks(input_folder, output_file):
    # Create a new workbook
    merged_wb = Workbook()

    # Iterate through each Excel file in the input folder
    for filename in os.listdir(input_folder):
        if filename.endswith(".xlsx"):
            filepath = os.path.join(input_folder, filename)
            # Load the workbook
            wb = load_workbook(filepath, read_only=True)
            # Assume each workbook has only one worksheet
            ws = wb.active
            # Copy data from the worksheet to the merged workbook
            for row in ws.iter_rows(values_only=True):
                merged_wb.active.append(row)

    # Save the merged workbook
    merged_wb.save(output_file)

# Example usage
input_folder = "G:\\Shared drives\\MILS -- MIN\\Statistics\\2024\\BibCounts\\milsbibs2024_todo"
output_file = "G:\\Shared drives\\MILS -- MIN\\Statistics\\2024\\BibCounts\\milsbibs2024_done\\BibCountCombined_2024.xlsx"
merge_workbooks(input_folder, output_file)





